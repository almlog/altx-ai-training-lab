# Copyright (c) 2026 Shunpei Suzuki (suzuki.shunpei@altx.co.jp), AltX Inc.
# Developed by Shunpei Suzuki <suzuki.shunpei@altx.co.jp>
#
"""Generates a realistic enterprise parameterized .xlsm sample procedure workbook."""

import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def generate_sample_xlsm(output_path: str):
    wb = openpyxl.Workbook()

    # Colors
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    accent_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    font_header = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Meiryo UI", size=9, bold=True)
    font_normal = Font(name="Meiryo UI", size=9)
    font_code = Font(name="Consolas", size=9, color="0F172A")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # -------------------------------------------------------------
    # Sheet 1: 環境パラメータ・承認情報
    # -------------------------------------------------------------
    ws_param = wb.active
    ws_param.title = "環境パラメータ・承認情報"
    ws_param.views.sheetView[0].showGridLines = True

    # Title
    ws_param["A1"] = "【株式会社AltX 本番システム移行・リリース作業計画書（パラメータ定義・承認印）】"
    ws_param["A1"].font = Font(name="Meiryo UI", size=13, bold=True, color="1E3A8A")

    # Section A: 承認情報
    ws_param["A3"] = "■ 1. 統括責任者 承認情報（Human Approval Record）"
    ws_param["A3"].font = font_bold

    approval_headers = ["承認ステータス", "承認責任者", "承認ID", "作業起票者", "承認年月日", "作業対象システム"]
    approval_values = [
        "承認済 (APPROVED)",
        "山田 太郎 (システム運用統括部長)",
        "APPR-20260906-ALTX-01",
        "鈴木 駿平 (AltX Inc.)",
        "2026-09-06",
        "基幹EC決済プラットフォーム (本番環境)",
    ]

    for col_idx, (h, v) in enumerate(zip(approval_headers, approval_values), start=1):
        cell_h = ws_param.cell(row=4, column=col_idx, value=h)
        cell_h.fill = sub_header_fill
        cell_h.font = font_header
        cell_h.alignment = Alignment(horizontal="center", vertical="center")
        cell_h.border = thin_border

        cell_v = ws_param.cell(row=5, column=col_idx, value=v)
        cell_v.font = font_bold if col_idx <= 2 else font_normal
        cell_v.alignment = Alignment(horizontal="center", vertical="center")
        cell_v.border = thin_border
        if col_idx == 1:
            cell_v.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            cell_v.font = Font(name="Meiryo UI", size=9, bold=True, color="15803D")

    # Section B: パラメータ定義テーブル
    ws_param["A7"] = "■ 2. 環境パラメータ定義テーブル（手順書内で ${変数名} として動的展開されます）"
    ws_param["A7"].font = font_bold

    param_headers = ["パラメータ変数名 (Key)", "本番設定値 (Value)", "パラメータ説明・注意事項", "設定検証結果"]
    for col_idx, h in enumerate(param_headers, start=1):
        c = ws_param.cell(row=8, column=col_idx, value=h)
        c.fill = header_fill
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    params_data = [
        ("${TARGET_HOST}", "db-prd-01.internal.altx.jp", "本番接続先ホストFQDN", "検証完了 (OK)"),
        ("${HEALTH_PORT}", "8080", "ヘルスチェックAPI監視ポート番号", "検証完了 (OK)"),
        ("${APP_VERSION}", "v2.1.0", "今回リリース対象の新パッケージバージョン", "検証完了 (OK)"),
        ("${BACKUP_DIR}", "/backup/20260906_release", "作業前完全バックアップ格納先ディレクトリ", "検証完了 (OK)"),
        ("${TARGET_TENANT}", "T100", "DB更新対象テナントID (限定更新)", "検証完了 (OK)"),
        ("${DB_USER}", "app_user", "本番MySQL接続実行ユーザー", "検証完了 (OK)"),
        ("${DB_NAME}", "app_db", "更新対象スキーマデータベース名", "検証完了 (OK)"),
    ]

    for row_idx, r in enumerate(params_data, start=9):
        for col_idx, val in enumerate(r, start=1):
            c = ws_param.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin_border
            c.font = font_code if col_idx <= 2 else font_normal
            if row_idx % 2 == 0:
                c.fill = zebra_fill
            if col_idx == 4:
                c.font = Font(name="Meiryo UI", size=9, bold=True, color="15803D")
                c.alignment = Alignment(horizontal="center")

    ws_param.column_dimensions["A"].width = 28
    ws_param.column_dimensions["B"].width = 38
    ws_param.column_dimensions["C"].width = 45
    ws_param.column_dimensions["D"].width = 22
    ws_param.column_dimensions["E"].width = 20
    ws_param.column_dimensions["F"].width = 38

    # -------------------------------------------------------------
    # Sheet 2: 本番作業手順書
    # -------------------------------------------------------------
    ws_proc = wb.create_sheet(title="本番作業手順書")
    ws_proc.views.sheetView[0].showGridLines = True

    ws_proc["A1"] = "【本番リリース実行手順書（HITMAN AIペアオペレーター Wチェック監視対象）】"
    ws_proc["A1"].font = Font(name="Meiryo UI", size=13, bold=True, color="1E3A8A")

    proc_headers = [
        "項番",
        "作業項目",
        "実行コマンド（パラメータ展開対象テンプレート）",
        "期待ログ・合否基準",
        "異常時・不整合時の自律分岐",
        "注意事項・リスク対策",
    ]

    for col_idx, h in enumerate(proc_headers, start=1):
        c = ws_proc.cell(row=3, column=col_idx, value=h)
        c.fill = header_fill
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    procedures = [
        (
            "1-1",
            "ディスク空き容量の事前確認",
            "df -h ${BACKUP_DIR}",
            "Avail >= 10GB",
            "異常時: 再投入ブロック (BLOCKED_RETRY)",
            "空き容量不足時はバックアップで枯渇するため本作業に進まないこと",
        ),
        (
            "1-2",
            "事前バックアップの取得（本作業）",
            "tar -czvf ${BACKUP_DIR}/app_${APP_VERSION}.tar.gz /var/www/app",
            "エラーなくアーカイブ作成",
            "異常時: ロールバック (R-1) へ分岐",
            "バックアップ完了までWebサービスを停止してはならない",
        ),
        (
            "2-1",
            "ロードバランサー切り離し確認",
            "curl -s -o /dev/null -w '%{http_code}' http://${TARGET_HOST}:${HEALTH_PORT}/health",
            "アクセス遮断確認",
            "異常時: エスカレーション (E-1) へ分岐",
            "外部トラフィックが確実に停止していることをログ検証",
        ),
        (
            "2-2",
            "Webサービスの停止（本作業）",
            "systemctl stop my-app.service",
            "Active: inactive (dead)",
            "異常時: ロールバック (R-1) へ分岐",
            "プロセスが完全に消滅したことを確認してからパッケージ同期へ進む",
        ),
        (
            "3-1",
            "環境設定ファイル（.env）退避",
            "cp -p /var/www/app/.env ${BACKUP_DIR}/.env.bak",
            ".env.bak 作成確認",
            "異常時: ロールバック (R-1) へ分岐",
            "パーミッションを保持したまま確実に退避する",
        ),
        (
            "3-2",
            "新バージョンの配置（本作業）",
            "rsync -avz --exclude='.env' /release/${APP_VERSION}/ /var/www/app/",
            "ファイル同期完了",
            "致命的クラッシュ検知時: ロールバック (R-1) へ分岐",
            ".env の誤上書きを避けるため除外オプション必須",
        ),
        (
            "3-3",
            "DB更新の事前確認（事前SELECT・SQL影響評価）",
            'mysql -u ${DB_USER} -p ${DB_NAME} -e "SELECT id, tenant_id, status, plan FROM users WHERE tenant_id = \'${TARGET_TENANT}\';"',
            "要件合致（MATCH）確認",
            "ロック競合・件数不一致時: エスカレーション (E-1) へ分岐",
            "テナントT100限定の更新であることをAIがSQL影響事前評価",
        ),
        (
            "3-4",
            "DB更新SQLの適用（本作業）",
            "mysql -u ${DB_USER} -p ${DB_NAME} < /release/${APP_VERSION}/db_update.sql",
            "affected rows 整合確認",
            "デッドロック時: エスカレーション (E-1) / 破損時: ロールバック (R-1)",
            "複数行が更新された場合は直ちにロールバック手順を発動",
        ),
        (
            "4-1",
            "新バージョンサービスの起動",
            "systemctl start my-app.service",
            "Active: active (running)",
            "異常時: ロールバック (R-1) へ分岐",
            "起動失敗時は直ちに journalctl -xe を確認",
        ),
        (
            "4-2",
            "正常性確認（リリース完了判定）",
            "curl -I http://${TARGET_HOST}:${HEALTH_PORT}/health",
            "HTTP/1.1 200 OK",
            "500応答時: ロールバック (R-1) へ分岐",
            "外部監視のアラートが解消されたことを確認",
        ),
    ]

    for row_idx, r in enumerate(procedures, start=4):
        for col_idx, val in enumerate(r, start=1):
            c = ws_proc.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin_border
            if col_idx == 1:
                c.font = font_bold
                c.alignment = Alignment(horizontal="center")
            elif col_idx == 3:
                c.font = font_code
            elif col_idx == 5:
                c.font = Font(name="Meiryo UI", size=8, bold=True, color="B45309")
                c.fill = accent_fill
            else:
                c.font = font_normal

            if row_idx % 2 != 0 and col_idx != 5:
                c.fill = zebra_fill

    ws_proc.column_dimensions["A"].width = 8
    ws_proc.column_dimensions["B"].width = 30
    ws_proc.column_dimensions["C"].width = 58
    ws_proc.column_dimensions["D"].width = 24
    ws_proc.column_dimensions["E"].width = 28
    ws_proc.column_dimensions["F"].width = 45

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"Sample Excel procedure workbook saved: {output_path}")


if __name__ == "__main__":
    target = os.path.join(
        os.path.dirname(__file__), "現場標準_Webアプリ本番リリース手順書_v2.1.xlsm"
    )
    generate_sample_xlsm(target)
