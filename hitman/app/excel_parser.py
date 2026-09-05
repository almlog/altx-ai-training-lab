# Copyright (c) 2026 Shunpei Suzuki (suzuki.shunpei@altx.co.jp), AltX Inc.
# Developed by Shunpei Suzuki <suzuki.shunpei@altx.co.jp>
#
"""Excel (.xlsm / .xlsx) SOP Parser for HITMAN.

Parses enterprise parameterized Excel procedures containing:
1. Approval Metadata Sheet (Author, Approver, Approval Date, Approval ID, Status)
2. Parameter Definition Sheet (Variables like ${TARGET_HOST}, ${DB_PORT}, ${APP_VERSION})
3. Execution Steps Sheet (Steps, Commands with parameters, Expected checks, Branch rules)
"""

import io
import re
from typing import Any, BinaryIO
import openpyxl


def parse_excel_sop(file_input: str | bytes | BinaryIO) -> dict[str, Any]:
    """Parse an .xlsm or .xlsx file and extract parameters, approval, and procedure steps.

    Args:
        file_input: File path (str), raw file bytes, or file-like BinaryIO stream.

    Returns:
        Dictionary containing:
            - status: 'success' | 'warning_unresolved' | 'error'
            - approval_metadata: dict (author, approver, approval_date, approval_id, is_approved)
            - parameters: dict of {KEY: value}
            - unresolved_parameters: list of variables that were not resolved
            - sop_database: dict matching HITMAN's SOP_DATABASE format
            - step_sequence: list of step IDs in execution order
            - branch_rules: dict of {step_id: {'on_failure': ..., 'on_mismatch': ...}}
            - message: summary message
    """
    try:
        if isinstance(file_input, bytes):
            wb = openpyxl.load_workbook(io.BytesIO(file_input), data_only=True)
        elif hasattr(file_input, "read"):
            wb = openpyxl.load_workbook(file_input, data_only=True)
        else:
            wb = openpyxl.load_workbook(file_input, data_only=True)
    except Exception as e:
        return {
            "status": "error",
            "message": f"Excelファイルの読み込みに失敗しました（.xlsm または .xlsx であることを確認してください）: {e}",
        }

    sheet_names = wb.sheetnames
    if not sheet_names:
        return {"status": "error", "message": "Excel内にシートが存在しません。"}

    # 1. シートの特定 (パラメータ・承認シート / 手順書本体シート)
    param_sheet = None
    procedure_sheet = None

    param_sheet_patterns = [r"パラメータ", r"parameter", r"環境変数", r"設定", r"config", r"承認", r"概要"]
    proc_sheet_patterns = [r"手順", r"procedure", r"作業", r"本番", r"execution", r"step", r"run"]

    for name in sheet_names:
        lower = name.lower()
        if not param_sheet and any(re.search(pat, lower, re.IGNORECASE) for pat in param_sheet_patterns):
            param_sheet = wb[name]
        elif not procedure_sheet and any(re.search(pat, lower, re.IGNORECASE) for pat in proc_sheet_patterns):
            procedure_sheet = wb[name]

    # フォールバック判定
    if not procedure_sheet:
        if len(sheet_names) == 1:
            procedure_sheet = wb[sheet_names[0]]
        else:
            # 最後のシートまたは2枚目を手順書とみなす
            procedure_sheet = wb[sheet_names[-1]]
            if not param_sheet and len(sheet_names) > 1:
                param_sheet = wb[sheet_names[0]]

    # 2. パラメータ＆承認情報の抽出
    parameters: dict[str, str] = {}
    approval_metadata = {
        "author": "作業担当者",
        "approver": "未承認（要確認）",
        "approval_date": "-",
        "approval_id": "-",
        "work_title": "本番システム移行・リリース作業",
        "is_approved": False,
    }

    if param_sheet:
        param_rows = [[str(c).strip() if c is not None else "" for c in r] for r in param_sheet.iter_rows(values_only=True)]
        
        # 2-A. 水平テーブル型 承認情報の探索 (Row N: ヘッダー, Row N+1: 値)
        approval_extracted = False
        for r_idx, row in enumerate(param_rows):
            # 承認関連のヘッダーキーワードが2つ以上含まれる行を探す
            approval_kw_count = sum(
                1 for c in row if any(k in c.replace(" ", "") for k in ["承認ステータス", "承認責任者", "承認者", "承認ID", "起票者", "作業起票者", "承認年月日", "承認日", "作業対象システム"])
            )
            if approval_kw_count >= 2 and r_idx + 1 < len(param_rows):
                val_row = param_rows[r_idx + 1]
                for c_idx, cell in enumerate(row):
                    clean_h = cell.replace(" ", "").replace("　", "")
                    val = val_row[c_idx] if c_idx < len(val_row) else ""
                    if not val:
                        continue
                    if any(k in clean_h for k in ["承認責任者", "承認者", "有識者承認"]):
                        approval_metadata["approver"] = val
                    elif any(k in clean_h for k in ["作業起票者", "起票者", "作成者", "担当者"]):
                        approval_metadata["author"] = val
                    elif any(k in clean_h for k in ["承認ID", "管理番号", "申請番号"]):
                        approval_metadata["approval_id"] = val
                    elif any(k in clean_h for k in ["承認年月日", "承認日", "確認日"]):
                        approval_metadata["approval_date"] = val
                    elif any(k in clean_h for k in ["作業対象システム", "作業件名", "手順書名", "件名"]):
                        approval_metadata["work_title"] = val
                    elif any(k in clean_h for k in ["承認ステータス", "ステータス"]):
                        if any(ok in val.upper() for ok in ["承認済", "APPROVED", "OK", "完了"]):
                            approval_metadata["is_approved"] = True
                approval_extracted = True
                break

        # 2-B. 垂直Key-Value型 または 単一セル型の承認情報探索 (未抽出項目のフォールバック)
        for r_idx, row in enumerate(param_rows):
            for c_idx, cell in enumerate(row):
                c_clean = cell.replace(" ", "").replace("　", "")
                val = row[c_idx + 1] if c_idx + 1 < len(row) else ""
                if not val or any(h in val for h in ["承認ID", "承認者", "起票者", "ステータス", "年月日"]):
                    continue
                if approval_metadata["approver"] == "未承認（要確認）" and any(k in c_clean for k in ["承認責任者", "承認者", "有識者承認"]):
                    approval_metadata["approver"] = val
                    approval_metadata["is_approved"] = True
                elif approval_metadata["author"] == "作業担当者" and any(k in c_clean for k in ["作業起票者", "起票者", "作成者"]):
                    approval_metadata["author"] = val
                elif approval_metadata["approval_id"] == "-" and any(k in c_clean for k in ["承認ID", "管理番号", "申請番号"]):
                    approval_metadata["approval_id"] = val
                elif approval_metadata["approval_date"] == "-" and any(k in c_clean for k in ["承認年月日", "承認日", "確認日"]):
                    approval_metadata["approval_date"] = val
                elif approval_metadata["work_title"] == "本番システム移行・リリース作業" and any(k in c_clean for k in ["作業対象システム", "作業件名", "手順書名"]):
                    approval_metadata["work_title"] = val

        # 承認ステータスが未判定でも承認責任者が有効に入力されていればTrueとする
        if approval_metadata["approver"] not in ["未承認（要確認）", "-", ""]:
            approval_metadata["is_approved"] = True

        # 2-C. パラメータテーブルの探索 (${KEY} または KEY と Value のペア)
        for row in param_rows:
            found_param = False
            for idx in range(len(row) - 1):
                raw_key = row[idx].strip()
                val = row[idx + 1].strip()
                if not raw_key or not val:
                    continue
                # パターン1: ${VAR} or {VAR}
                m1 = re.match(r"^\$\{([A-Za-z0-9_]+)\}$", raw_key)
                if not m1:
                    m1 = re.match(r"^\{([A-Za-z0-9_]+)\}$", raw_key)
                # パターン2: 大文字スネークケース (TARGET_HOST など、英字始まり3文字以上、かつ値ではない)
                if not m1 and re.match(r"^[A-Z][A-Z0-9_]{2,}$", raw_key):
                    if raw_key not in ["KEY", "PARAM", "PARAMETER", "NAME", "VALUE", "DESC", "STATUS", "DATE", "TRUE", "FALSE"]:
                        parameters[raw_key] = val
                        found_param = True
                        break
                elif m1:
                    k = m1.group(1).upper()
                    if k not in ["KEY", "PARAM", "PARAMETER", "NAME", "変数名", "項目名", "キー"]:
                        parameters[k] = val
                        found_param = True
                        break
            if found_param:
                continue

    # 3. 手順書本体シートの解析
    if not procedure_sheet:
        return {"status": "error", "message": "手順書シートが見つかりませんでした。"}

    header_row_idx = None
    col_map = {
        "step_id": -1,
        "title": -1,
        "command": -1,
        "expected": -1,
        "branch": -1,
        "cautions": -1,
    }

    rows = list(procedure_sheet.iter_rows(values_only=True))
    for r_idx, row in enumerate(rows):
        if not row:
            continue
        row_str = [str(c).strip().replace(" ", "").replace("　", "") if c is not None else "" for c in row]
        has_id = any(h in cell for cell in row_str for h in ["項番", "ステップ", "step", "no.", "no"])
        has_cmd = any(h in cell for cell in row_str for h in ["コマンド", "command", "実行内容", "投入"])
        if has_id and (has_cmd or any("作業" in cell for cell in row_str)):
            header_row_idx = r_idx
            for c_idx, cell in enumerate(row_str):
                if any(h in cell for h in ["項番", "ステップ", "step", "no"]):
                    col_map["step_id"] = c_idx
                elif any(h in cell for h in ["作業項目", "作業内容", "タイトル", "title", "タスク"]):
                    col_map["title"] = c_idx
                elif any(h in cell for h in ["コマンド", "command", "投入コマンド", "実行コマンド"]):
                    col_map["command"] = c_idx
                elif any(h in cell for h in ["期待結果", "期待ログ", "判定基準", "確認項目", "expected"]):
                    col_map["expected"] = c_idx
                elif any(h in cell for h in ["分岐", "異常時", "合否分岐", "branch"]):
                    col_map["branch"] = c_idx
                elif any(h in cell for h in ["注意", "備考", "caution", "補足"]):
                    col_map["cautions"] = c_idx
            break

    if header_row_idx is None or col_map["step_id"] == -1:
        col_map = {"step_id": 0, "title": 1, "command": 2, "expected": 3, "branch": 4, "cautions": 5}
        header_row_idx = 0

    parsed_steps = []
    for r_idx in range(header_row_idx + 1, len(rows)):
        row = rows[r_idx]
        if not row or all(c is None for c in row):
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]

        step_id = cells[col_map["step_id"]] if col_map["step_id"] < len(cells) else ""
        if not step_id or step_id.lower() in ["項番", "step", "no"]:
            continue

        title = cells[col_map["title"]] if col_map["title"] != -1 and col_map["title"] < len(cells) else f"ステップ {step_id}"
        cmd = cells[col_map["command"]] if col_map["command"] != -1 and col_map["command"] < len(cells) else ""
        expected = cells[col_map["expected"]] if col_map["expected"] != -1 and col_map["expected"] < len(cells) else "正常終了すること"
        branch = cells[col_map["branch"]] if col_map["branch"] != -1 and col_map["branch"] < len(cells) else ""
        cautions = cells[col_map["cautions"]] if col_map["cautions"] != -1 and col_map["cautions"] < len(cells) else "引数・オプションを確認して慎重に投入すること"

        parsed_steps.append({
            "raw_step_id": step_id,
            "title": title or f"ステップ {step_id}",
            "command_template": cmd,
            "expected_check": expected or "正常終了すること",
            "branch_raw": branch,
            "cautions": cautions,
        })

    if not parsed_steps:
        return {"status": "error", "message": "手順書シートから手順ステップ行を抽出できませんでした。"}

    # 4. パラメータの展開（Parameter Substitution）と未解決パラメータの検査
    def substitute_params(text: str) -> str:
        if not text:
            return ""
        result = text
        for k, v in parameters.items():
            result = re.sub(rf"\$\{{{k}\}}", v, result, flags=re.IGNORECASE)
            result = re.sub(rf"\{{{k}\}}", v, result, flags=re.IGNORECASE)
        return result

    unresolved_set = set()
    for s in parsed_steps:
        expanded_cmd = substitute_params(s["command_template"])
        expanded_exp = substitute_params(s["expected_check"])
        s["command"] = expanded_cmd
        s["expected_check"] = expanded_exp

        # 未解決変数の探索 (${...} or {...})
        for match in re.finditer(r"(\$\{([A-Za-z0-9_]+)\})", expanded_cmd):
            unresolved_set.add(match.group(2).upper())
        for match in re.finditer(r"(\$\{([A-Za-z0-9_]+)\})", expanded_exp):
            unresolved_set.add(match.group(2).upper())

    # 5. 分岐ルールの構造化
    branch_rules: dict[str, dict[str, str]] = {}
    for s in parsed_steps:
        sid = s["raw_step_id"]
        raw_b = s.get("branch_raw", "")
        rule = {}
        if re.search(r"R-1|ロールバック|切り戻し|巻き戻し", raw_b, re.IGNORECASE):
            rule["on_failure"] = "R-1"
        if re.search(r"E-1|エスカレーション|協議|ゲート", raw_b, re.IGNORECASE):
            rule["on_mismatch"] = "E-1"
        branch_rules[sid] = rule

    # 6. HITMAN SOP_DATABASE への構造化マッピング
    sop_db = {}
    step_seq = []

    for idx, s in enumerate(parsed_steps):
        raw_id = s["raw_step_id"]
        title = s["title"]
        cmd = s["command"]
        exp = s["expected_check"]
        cau = s["cautions"]

        if "-" in raw_id:
            parent_part = raw_id.split("-")[0]
            parent_key = int(parent_part) if parent_part.isdigit() else parent_part
        else:
            parent_key = int(raw_id) if raw_id.isdigit() else (idx + 1)
            raw_id = f"{parent_key}-1"

        if parent_key not in sop_db:
            sop_db[parent_key] = {
                "step_id": str(parent_key),
                "title": f"ステップ {parent_key}: {title}",
                "objective": title,
                "command": cmd,
                "expected_check": exp,
                "cautions": cau,
                "sub_steps": {},
            }

        sop_db[parent_key]["sub_steps"][raw_id] = {
            "title": f"ステップ {raw_id}: {title}",
            "objective": title,
            "command": cmd,
            "expected_check": exp,
            "cautions": cau,
            "branch_rule": branch_rules.get(s["raw_step_id"], {}),
        }
        step_seq.append(raw_id)

    # 必須の標準エスカレーションとロールバックが未定義なら安全のため補完
    if "E" not in sop_db:
        sop_db["E"] = {
            "step_id": "E",
            "title": "エスカレーション対応（GO/NOGO判定ゲート）",
            "objective": "想定外事象・不整合発生時の有識者・上長協議ゲート",
            "command": "# [エスカレーション・ゲート] UIフォームより対応結果・GO/NOGO・判断根拠を入力",
            "expected_check": "判断根拠および承認の確認",
            "cautions": "判断根拠（こんきょ）なしには再開できません",
            "sub_steps": {
                "E-1": {
                    "title": "ステップ E-1: エスカレーション協議とGO/NOGO・根拠確定",
                    "objective": "有識者協議と特別対応判定",
                    "command": "# [エスカレーション・ゲート] UIフォームより入力",
                    "expected_check": "判断根拠および承認の確認",
                    "cautions": "判断根拠なしには進行不可",
                }
            },
        }

    if "R" not in sop_db:
        sop_db["R"] = {
            "step_id": "R",
            "title": "ロールバック手順（異常復旧）",
            "objective": "旧バージョンへの切り戻しと健全性復旧",
            "command": "tar -xzvf /backup/app_*.tar.gz -C / && cp -p /backup/.env.bak /var/www/app/.env",
            "expected_check": "旧ファイルの完全復元と健全稼働確認",
            "cautions": "現行の破損ファイルが残らないよう確認",
            "sub_steps": {
                "R-1": {
                    "title": "ステップ R-1: バックアップからの旧バージョン復元",
                    "objective": "事前バックアップの展開と設定ファイル復元",
                    "command": "tar -xzvf /backup/app_*.tar.gz -C / && cp -p /backup/.env.bak /var/www/app/.env",
                    "expected_check": "旧ファイルの完全復元",
                    "cautions": "現行破損ファイルの混入防止",
                },
                "R-2": {
                    "title": "ステップ R-2: サービス再起動と健全性確認",
                    "objective": "サービスの再起動とHTTP 200復旧確認",
                    "command": "systemctl restart my-app.service && curl -I http://localhost:8080/health",
                    "expected_check": "HTTP/1.1 200 OK 復帰",
                    "cautions": "復旧完了後、障害報告書の証跡を保全すること",
                },
            },
        }

    status = "warning_unresolved" if unresolved_set else "success"
    unresolved_list = sorted(list(unresolved_set))

    return {
        "status": status,
        "sheet_names": sheet_names,
        "approval_metadata": approval_metadata,
        "parameters": parameters,
        "unresolved_parameters": unresolved_list,
        "imported_steps_count": len(step_seq),
        "sop_database": sop_db,
        "step_sequence": step_seq,
        "branch_rules": branch_rules,
        "message": (
            f"【.xlsm 手順書解析成功】{len(step_seq)} 件の手順ステップと {len(parameters)} 個の環境パラメータを抽出しました。"
            + (f" ⚠️ 未解決のパラメータがあります: {unresolved_list}" if unresolved_list else " 全パラメータが正常に展開されました。")
        ),
    }
